"""Extract golden fixtures from the private source workbook.

The workbook is NEVER committed. Set D365_ESTIMATE_XLSX to its path and run:

    python tests/golden/extract_fixtures.py

Writes tests/golden/fixtures/local/workbook.json (gitignored) with the
computed values the engine must reproduce: per-month license counts,
included-storage entitlements, storage-needed subtotals, storage overage
costs, and the Report category/grand-total rows.

Requires the workbook to have been opened and saved by Excel so cached
formula values exist (openpyxl data_only=True reads the cache).
"""
import json
import os
import sys
from pathlib import Path

import openpyxl

MONTHS = 60
FIRST_MONTH_COL = 8  # column H = month 1 on the Worksheet sheet


def row_values(ws, row: int) -> list:
    out = []
    for m in range(MONTHS):
        v = ws.cell(row=row, column=FIRST_MONTH_COL + m).value
        out.append(v if isinstance(v, (int, float)) else None)
    return out


def main() -> int:
    path = os.environ.get("D365_ESTIMATE_XLSX")
    if not path or not Path(path).exists():
        print("Set D365_ESTIMATE_XLSX to the workbook path.", file=sys.stderr)
        return 1

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Worksheet"]
    report = wb["Report"]

    fixture = {
        "source": Path(path).name,
        "licenseCounts": {
            "erpPremium": row_values(ws, 5),
            "erpFull": row_values(ws, 6),
            "cePremium": row_values(ws, 7),
            "ceEnterprise": row_values(ws, 8),
            "csProfessional": row_values(ws, 9),
            "attach": row_values(ws, 10),
            "activity": row_values(ws, 11),
            "teamMember": row_values(ws, 12),
            "device": row_values(ws, 13),
        },
        # Included storage: tenant + CAL accrual pairs (rows 177..186)
        "includedStorage": {
            "fscmData": [
                (a or 0) + (b or 0)
                for a, b in zip(row_values(ws, 177), row_values(ws, 178))
            ],
            "fscmFile": [
                (a or 0) + (b or 0)
                for a, b in zip(row_values(ws, 179), row_values(ws, 180))
            ],
            "dvData": [
                (a or 0) + (b or 0)
                for a, b in zip(row_values(ws, 181), row_values(ws, 182))
            ],
            "dvFile": [
                (a or 0) + (b or 0)
                for a, b in zip(row_values(ws, 183), row_values(ws, 184))
            ],
        },
        # Total storage needed (array-formula subtotals, rows 187..190)
        "neededStorage": {
            "fscmData": row_values(ws, 187),
            "fscmFile": row_values(ws, 188),
            "dvData": row_values(ws, 189),
            "dvFile": row_values(ws, 190),
        },
        # Monthly overage cost rows (191..194)
        "storageCost": {
            "fscmData": row_values(ws, 191),
            "fscmFile": row_values(ws, 192),
            "dvData": row_values(ws, 193),
            "dvFile": row_values(ws, 194),
        },
        # Overage unit prices (H18..H21)
        "overagePrices": {
            "fscmData": ws.cell(row=18, column=8).value,
            "fscmFile": ws.cell(row=19, column=8).value,
            "dvData": ws.cell(row=20, column=8).value,
            "dvFile": ws.cell(row=21, column=8).value,
        },
        # Report grand total row 21, months in columns B..BI
        "reportGrandTotal": [
            v if isinstance(v := report.cell(row=21, column=2 + m).value, (int, float)) else None
            for m in range(MONTHS)
        ],
    }

    out_dir = Path(__file__).parent / "fixtures" / "local"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / "workbook.json"
    out_path.write_text(json.dumps(fixture, indent=1))
    print(f"Wrote {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
